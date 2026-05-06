# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.chart import PieChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT = r"d:\My_Budget_v2.xlsx"
RATE = 43.81
INVEST = 100_000   # $ initial investment

# ── helpers ───────────────────────────────────────────────────────────────────
def S(style='thin'): return Side(style=style)
def B(t=0,b=0,l=0,r=0):
    return Border(top=S() if t else None, bottom=S() if b else None,
                  left=S() if l else None,  right=S() if r else None)
def F(c): return PatternFill('solid', fgColor=c)
def fmt_usd(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = '"$"#,##0.00'
    return c

HDR_DARK  = F('1F4E79')
HDR_MID   = F('2E75B6')
WHITE_FNT = Font(bold=True, color='FFFFFF')
BOLD      = Font(bold=True)

def hdr(ws, row, col, val, fill_=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = WHITE_FNT if fill_ else BOLD
    if fill_: c.fill = fill_
    c.alignment = Alignment(horizontal='center', wrap_text=True)
    c.border = B(1,1,1,1)
    return c

def row_border(ws, row, cols):
    for col in range(1, cols+1):
        ws.cell(row=row, column=col).border = B(1,1,1,1)

# ── Sheet 1: Постійні витрати ─────────────────────────────────────────────────
def sheet_costs(wb):
    ws = wb.active
    ws.title = 'Постійні витрати'
    for col,w in zip('ABCDEFG',[22,16,22,14,16,14,10]):
        ws.column_dimensions[col].width = w

    for i,h in enumerate(['Команда','Тип оплати','Технологія',
                           'Сума оплати','Час роботи, міс','Загальна сума','Джерело'],1):
        hdr(ws,2,i,h,HDR_DARK)

    team = [
        ('Фронт-енд дев','Fixed Salary','JavaScript/HTML/CSS',2800,3,8400,'DOU'),
        ('Бекенд дев',   'Fixed Salary','Django (Python)',     2400,3,7200,'DOU'),
        ('QA',           'Fixed Salary','Manual testing',      1600,2,3200,'DOU'),
        ('UX/UI Designer','Hourly Payment','WEB Design',       1600,1,1600,'DOU'),
    ]
    for r,row in enumerate(team,3):
        for c,v in enumerate(row,1):
            cell = ws.cell(row=r,column=c,value=v)
            cell.border = B(1,1,1,1)
            cell.alignment = Alignment(horizontal='center')
            if c in(4,6): cell.number_format='"$"#,##0.00'

    # Постійні витрати
    for i,h in enumerate(['Постійні витрати','Сума','Джерело'],1):
        hdr(ws,8,i,h,HDR_MID)
    rec = [('Хостинг/Server/Cloud',50,'https://thehost.ua/'),
           ('Domain name',1.25,'imena.ua'),
           ('Всього',51.25,None)]
    for r,(n,a,s) in enumerate(rec,9):
        ws.cell(row=r,column=1,value=n).border=B(1,1,1,1)
        c=fmt_usd(ws,r,2,a); c.border=B(1,1,1,1)
        if n=='Всього': ws.cell(row=r,column=1).font=BOLD; c.font=BOLD
        if s: ws.cell(row=r,column=3,value=s).border=B(1,1,1,1)

    # Калькуляція стартових затрат
    ws.cell(row=13,column=1,value='Калькуляція стартових затрат').fill=HDR_MID
    ws.cell(row=13,column=1).font=WHITE_FNT
    ws.cell(row=13,column=3,value='Примітки').fill=HDR_MID
    ws.cell(row=13,column=3).font=WHITE_FNT
    startup=[('Фронт-енд дев',8400,''),('Бекенд дев',7200,''),
             ('QA',3200,''),('UX/UI Designer',1600,''),
             ('Маркетолог',5000,'Project-based Payment'),('Всього',25400,None)]
    for r,(n,a,note) in enumerate(startup,14):
        ws.cell(row=r,column=1,value=n).border=B(1,1,1,1)
        c=fmt_usd(ws,r,2,a); c.border=B(1,1,1,1)
        if n=='Всього': ws.cell(row=r,column=1).font=BOLD; c.font=BOLD
        if note is not None:
            ws.cell(row=r,column=3,value=note).border=B(1,1,1,1)

# ── Sheet 2: Джерела доходу ──────────────────────────────────────────────────
def sheet_revenue(wb):
    ws = wb.create_sheet('Джерела доходу')
    for col,w in zip('ABC',[32,14,22]):
        ws.column_dimensions[col].width=w
    for i,h in enumerate(['Джерело','Вартість','Потенційна аудиторія'],1):
        hdr(ws,1,i,h,HDR_DARK)

    ws.merge_cells('A2:C2')
    hdr(ws,2,1,'Основні',HDR_MID)
    main=[('Підписка',20,20000),('Продаж API для інших платформ',150,1000)]
    for r,(n,p,a) in enumerate(main,3):
        ws.cell(row=r,column=1,value=n).border=B(1,1,1,1)
        c=fmt_usd(ws,r,2,p); c.border=B(1,1,1,1)
        ws.cell(row=r,column=3,value=a).border=B(1,1,1,1)

    ws.merge_cells('A5:C5')
    hdr(ws,5,1,'Додаткові',HDR_MID)
    for r,n in enumerate(['Рекламна інтеграція','Мерч',
                          'Співпраця з брендами(агентствами)'],6):
        ws.cell(row=r,column=1,value=n).border=B(1,1,1,1)
        ws.cell(row=r,column=2).border=B(1,1,1,1)
        ws.cell(row=r,column=3).border=B(1,1,1,1)

# ── Sheet 3: Маркетинг + pie chart ───────────────────────────────────────────
def sheet_marketing(wb):
    ws = wb.create_sheet('Маркетинг')
    for col,w in zip('ABCD',[24,12,8,12]):
        ws.column_dimensions[col].width=w

    for i,h in enumerate(['Канали просування','Price','Qt','Sum'],1):
        c=ws.cell(row=1,column=i,value=h); c.font=BOLD; c.border=B(b=1)

    ws.cell(row=2,column=1,value='Соціальні мережі (SMM)').font=BOLD
    smm=[('Instagram',500,5,2500),('TikTok',50,20,1000),
         ('YouTube',None,None,None),('LinkedIn',None,None,None)]
    for r,(n,p,q,s) in enumerate(smm,3):
        ws.cell(row=r,column=1,value=n).alignment=Alignment(horizontal='right')
        if p:
            fmt_usd(ws,r,2,p); ws.cell(row=r,column=3,value=q)
            fmt_usd(ws,r,4,s)

    ws.cell(row=7,column=1,value='Контекстна реклама')
    fmt_usd(ws,7,4,2500)
    ws.cell(row=8,column=1,value='SEO')
    fmt_usd(ws,8,4,3000)
    c=fmt_usd(ws,10,4,9000); c.font=BOLD

    # chart data
    chart_data=[('SMM',3500),('Контекстна реклама',2500),('SEO',3000)]
    for i,(n,v) in enumerate(chart_data,12):
        ws.cell(row=i,column=1,value=n)
        ws.cell(row=i,column=2,value=v)

    chart=PieChart(); chart.title='Канали просування'; chart.style=10
    data=Reference(ws,min_col=2,min_row=12,max_row=14)
    cats=Reference(ws,min_col=1,min_row=12,max_row=14)
    chart.add_data(data); chart.set_categories(cats)
    colors=['4472C4','FF0000','FFC000']
    for idx,color in enumerate(colors):
        pt=DataPoint(idx=idx); pt.graphicalProperties.solidFill=color
        chart.series[0].dPt.append(pt)
    chart.width=14; chart.height=10
    ws.add_chart(chart,'F1')

# ── Sheet 4: Total ────────────────────────────────────────────────────────────
def sheet_total(wb):
    ws = wb.create_sheet('Total')

    # column widths
    widths = [8,12,12,14,14, 10,10,10,14,14,16, 12,14,12,16]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_col(i)].width=w

    # header row 1
    groups = [('Місяць',1),('К-сть користувачів',4),('Затрати',6),('Дохід',4)]
    col=1
    for grp,span in groups:
        if span>1:
            ws.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+span-1)
        c=ws.cell(row=1,column=col,value=grp)
        c.font=WHITE_FNT; c.fill=HDR_DARK
        c.alignment=Alignment(horizontal='center'); c.border=B(1,1,1,1)
        col+=span

    # header row 2
    sub = ['Місяць',
           'Контекст','SMM','Орган. трафік','Всього',
           'Контекст','SMM','SEO','Постійні','Всього затрат','Накоп. витрат',
           'Маркетинг','Прибуток','Чистий приб.','Накоп. доходу']
    for i,h in enumerate(sub,1):
        c=ws.cell(row=2,column=i,value=h)
        c.font=WHITE_FNT; c.fill=HDR_MID
        c.alignment=Alignment(horizontal='center',wrap_text=True)
        c.border=B(1,1,1,1)
    ws.row_dimensions[2].height=30

    # --- calculate monthly data ---
    cum_cost=0; cum_profit=0
    INVEST_USD=INVEST
    breakeven_row=None; payback_row=None

    for m in range(1,37):
        # users
        ctx = m*10
        smm = m*70/3
        org = max(0, m-6)
        total_u = ctx+smm+org

        # costs by phase
        if m<=6:    c_ctx,c_smm,c_seo=2500,3500,3000
        elif m<=12: c_ctx,c_smm,c_seo=2500,3500,1500
        elif m<=24: c_ctx,c_smm,c_seo=2500,3500,1000
        else:       c_ctx,c_smm,c_seo=500,500,200
        fixed=51.25
        total_c=c_ctx+c_smm+c_seo+fixed
        cum_cost+=total_c

        # revenue (ARPU=$20 subscription)
        revenue=total_u*20
        profit=revenue-total_c
        net=profit*0.95
        cum_profit+=net

        row=m+2
        vals=[m, ctx, smm, org, total_u,
              c_ctx,c_smm,c_seo,fixed,total_c,cum_cost,
              revenue,profit,net,cum_profit]
        for col,v in enumerate(vals,1):
            c=ws.cell(row=row,column=col,value=round(v,2))
            c.border=B(1,1,1,1)
            c.alignment=Alignment(horizontal='right')
            if col in(11,15): c.number_format='#,##0.00'

        # detect break-even (first month profit > 0)
        if breakeven_row is None and profit>0:
            breakeven_row=row
        # detect payback (cumulative profit >= investment)
        if payback_row is None and cum_profit>=INVEST_USD:
            payback_row=row

    # highlight break-even (yellow)
    if breakeven_row:
        for col in range(1,16):
            ws.cell(row=breakeven_row,column=col).fill=F('FFFF00')
        ws.cell(row=breakeven_row,column=16,value='← Точка беззбитковості').font=Font(bold=True,color='B8860B')

    # highlight payback (green)
    if payback_row:
        for col in range(1,16):
            ws.cell(row=payback_row,column=col).fill=F('00B050')
            ws.cell(row=payback_row,column=col).font=Font(color='FFFFFF',bold=True)
        ws.cell(row=payback_row,column=16,value='← Точка окупності').font=Font(bold=True,color='00B050')

    # totals
    last=38
    ws.cell(row=last,column=11,value=round(cum_cost,2)).font=BOLD
    ws.cell(row=last,column=15,value=round(cum_profit,2)).font=BOLD

    data_min_row = 3
    data_max_row = 38
    cats = Reference(ws, min_col=1, min_row=data_min_row, max_row=data_max_row)

    # ── Chart 1: Накопичені витрати vs Накопичені доходи ────────────────────
    ch1 = LineChart()
    ch1.title  = "Накопичені витрати та доходи"
    ch1.style  = 10
    ch1.y_axis.title = "Сума ($)"
    ch1.x_axis.title = "Місяць"
    ch1.height = 14
    ch1.width  = 22

    ref_cum_cost = Reference(ws, min_col=11, min_row=data_min_row, max_row=data_max_row)
    ch1.add_data(ref_cum_cost)
    ch1.series[0].title = SeriesLabel(v="Накоп. витрати")
    ch1.series[0].graphicalProperties.line.solidFill = "FF0000"
    ch1.series[0].graphicalProperties.line.width     = 20000
    ch1.series[0].smooth = True

    ref_cum_inc = Reference(ws, min_col=15, min_row=data_min_row, max_row=data_max_row)
    ch1.add_data(ref_cum_inc)
    ch1.series[1].title = SeriesLabel(v="Накоп. доходи")
    ch1.series[1].graphicalProperties.line.solidFill = "4472C4"
    ch1.series[1].graphicalProperties.line.width     = 20000
    ch1.series[1].smooth = True

    ch1.set_categories(cats)
    ws.add_chart(ch1, "A41")   # ← під таблицею зліва

    # ── Chart 2: Щомісячні витрати vs Чистий прибуток ───────────────────────
    ch2 = LineChart()
    ch2.title  = "Щомісячні витрати та прибуток"
    ch2.style  = 10
    ch2.y_axis.title = "Сума ($)"
    ch2.x_axis.title = "Місяць"
    ch2.height = 14
    ch2.width  = 22

    ref_mon_cost = Reference(ws, min_col=10, min_row=data_min_row, max_row=data_max_row)
    ch2.add_data(ref_mon_cost)
    ch2.series[0].title = SeriesLabel(v="Всього затрат")
    ch2.series[0].graphicalProperties.line.solidFill = "4472C4"
    ch2.series[0].graphicalProperties.line.width     = 20000
    ch2.series[0].smooth = True

    ref_mon_net = Reference(ws, min_col=14, min_row=data_min_row, max_row=data_max_row)
    ch2.add_data(ref_mon_net)
    ch2.series[1].title = SeriesLabel(v="Чистий прибуток")
    ch2.series[1].graphicalProperties.line.solidFill = "FF0000"
    ch2.series[1].graphicalProperties.line.width     = 20000
    ch2.series[1].smooth = True

    ch2.set_categories(cats)
    ws.add_chart(ch2, "M41")   # ← під таблицею справа

    be_m = breakeven_row-2 if breakeven_row else 'N/A'
    pb_m = payback_row-2   if payback_row   else 'N/A'
    print(f"Break-even: місяць {be_m} (жовтий)")
    print(f"Payback:    місяць {pb_m} (зелений)")
    print(f"Графік додано на аркуш Total (комірка R2)")

# ── Sheet 5: Фінансові показники ──────────────────────────────────────────────
def sheet_metrics(wb):
    ws=wb.create_sheet('Фінансові показники')
    for col,w in zip('ABCDEF',[18,16,18,4,16,12]):
        ws.column_dimensions[col].width=w

    metrics=[
        ('ROI','Інвестиції','Чистий прибуток','','Поточний курс',RATE),
        ('160%',  100000,   160000,'','',''),
        ('','','','','',''),
        ('CAC','Витрати','К-сть користувачів','','',''),
        (37.51, 100000, 2666,'','',''),
        ('1 643,29 грн.','','','','',''),
        ('','','','','',''),
        ('ARPU','Критична к-сть користувачів','','','',''),
        (12.50, 240,'','','',''),
        ('547,63 грн.','','','','',''),
        ('','','','','',''),
        ('LTV','Час користування','Дохід в міс','','',''),
        (300, 24, 12.50,'','',''),
        ('13 143,00 грн.','','','','',''),
        ('','','','','',''),
        ('','дол. США','грн','','',''),
        ('Загальні витрати',  100000, f'{100000*RATE:,.0f} грн.','','',''),
        ('Загальний прибуток',160000, f'{160000*RATE:,.0f} грн.','','',''),
        ('ROI','160%','','','',''),
        ('CAC',37.51,f'1 643,29 грн.','','',''),
        ('ARPU',12.50,'547,63 грн.','','',''),
        ('LTV',300,'13 143,00 грн.','','',''),
    ]

    bold_rows={0,3,7,11,15}
    for r,row in enumerate(metrics,1):
        for c,v in enumerate(row,1):
            cell=ws.cell(row=r,column=c,value=v)
            if r-1 in bold_rows and c<=3:
                cell.font=WHITE_FNT
                cell.fill=HDR_MID
            if c in(2,3) and isinstance(v,(int,float)) and r not in(1,):
                cell.number_format='"$"#,##0.00'

def get_col(n):
    return openpyxl.utils.get_column_letter(n)

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    wb=openpyxl.Workbook()
    sheet_costs(wb)
    sheet_revenue(wb)
    sheet_marketing(wb)
    sheet_total(wb)
    sheet_metrics(wb)
    wb.save(OUT)
    print(f"Saved: {OUT}")

if __name__=='__main__':
    main()
